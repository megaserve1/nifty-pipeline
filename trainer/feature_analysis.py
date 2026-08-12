"""trainer/feature_analysis.py -- the FEATURE ANALYSIS workbook, one file per run.

WHAT IT PRODUCES
    two sheets, generated from scratch in the shape of `feature analysis.xlsx`:

      "Incorrect Classification SHAP"   42 rows -- EVERY wrong (true -> predicted) pair
      "Correct Classification SHAP"      7 rows -- the diagonal

    7 classes gives 49 pairs and all 49 get a row, whether the model produced them or not. a pair
    that never happened shows count 0. dropping it would make "never happened" and "not looked
    at" the same thing, and the sheet would change shape between runs.

    each row gets:
      test # of Classifications   how many test rows fell in that cell
      one column PER FEATURE      the mean SIGNED shap of that feature, TOWARD THE PREDICTED CLASS

THE FEATURE COLUMNS ARE THE RUN'S OWN, NOT A FIXED LIST.
    the template ships 25 hand-written group headings (trend_direction, vol_phase, ...). those
    describe the FULL feature set; a given run trains on whatever its dataset version carried --
    v10 has 22 columns, 10 of those headings have no feature at all, and 3 of its columns have no
    heading. writing the run's actual features means the sheet always matches the model that
    produced it, with nothing silently blank and nothing silently dropped.

WHY SIGNED, AND TOWARD THE PREDICTED CLASS
    the template ships values like 0.72 / -0.54 / 0.23 -- negatives, so it is signed shap, not
    mean(|shap|). and for a cell "true was A, model said B" the question is what pushed it to say
    B, so the class axis is B. same convention as shap_explain's "WHY A became B" table.

    signed values CANCEL across rows (measured: stress_signal_label_combined cancels 99% on the
    full test set). that is intended -- a feature that pushes both ways on a given mistake has no
    consistent story, and a near-zero cell is the honest answer.

NO TEMPLATE FILE. the workbook we were given is a FORMAT, not an input. reading it back in
    dragged along four sheets this does not implement (Drawdowns, Losing Streak, FC Severity, FC
    Wobble), its prose columns, and its row list -- which is 44 rows, not 49. so the rows are
    generated here and the severity comes from configs/severity_7class.json, the same file the
    trainer's objective reads.

RUN
    final_venv/bin/python trainer/feature_analysis.py --model_task_id <training task id>
"""
import argparse
import pathlib
import sys
import time

import numpy as np
import pandas as pd

_here = pathlib.Path(__file__).resolve()
sys.path.insert(0, str(_here.parent.parent))
sys.path.insert(0, str(_here.parent.parent / "scripts"))   # predict.prepare
import config as C                                    # noqa: E402
from trainer.shap_logic import compute_shap           # noqa: E402
from trainer.streaks import (                          # noqa: E402
    cell_rows, TIER_ORDER, DEF_DD, DEF_PERIOD, DEF_LS)

SHEET_BAD = "Incorrect Classification SHAP"
SHEET_OK = "Correct Classification SHAP"
COL_TRUE, COL_PRED = "True Action", "Predicted Action"
COL_SEV, COL_N, COL_W = "Severity Score", "test # of Classifications", "class weight (true class)"
COL_CHG, COL_RAT = "Change", "Rationale"
SHEET_DD = "Classification Drawdowns"
SHEET_LS = "Losing Streak"
SHEET_CM = "Confusion Matrix"
SHEET_SUP = "Cell Support Data"


def _order(recs):
    """errors first, grouped by tier and severity, then the 7 correct rows."""
    return sorted(recs, key=lambda r: (r["tier"] is None,
                                       TIER_ORDER.index(r["tier"]) if r["tier"] else 0,
                                       r["severity"], r["true"]))


def dd_sheet(recs, prev=None) -> pd.DataFrame:
    rows=[{"Tier": r["tier"], COL_TRUE: r["true"], COL_PRED: r["pred"],
           "Max DD1 of classifications": r["max_dd1"], "DD1 Period": r["dd1_period"],
           "Max DD2 of classifications": r["max_dd2"], "DD2 Period": r["dd2_period"],
           "Max DD3 of classifications": r["max_dd3"], "DD3 Period": r["dd3_period"],
           COL_CHG: (f"{r['max_dd1']}-{prev[(r['true'],r['pred'])]}"
                     if prev and (r["true"],r["pred"]) in prev and r["max_dd1"] is not None
                     else ""),
           COL_RAT: ""} for r in _order(recs)]
    return pd.DataFrame(rows)


def ls_sheet(recs) -> pd.DataFrame:
    rows=[{"Tier": r["tier"], COL_TRUE: r["true"], COL_PRED: r["pred"],
           COL_SEV: r["severity"],
           "Max Losing Streak1": r["max_streak1"], "Losing Streak1 Period": r["streak1_period"],
           "Max Losing Streak2": r["max_streak2"], "Losing Streak2 Period": r["streak2_period"],
           "Max Losing Streak3": r["max_streak3"], "Losing Streak3 Period": r["streak3_period"],
           COL_CHG: "", COL_RAT: ""} for r in _order(recs)]
    return pd.DataFrame(rows)


def cm_sheet(y_true, y_pred, classes) -> pd.DataFrame:
    """rows = TRUE, columns = PREDICTED. the diagonal is coloured green afterwards."""
    m = pd.crosstab(pd.Series(y_true, name="TRUE"), pd.Series(y_pred, name="PREDICTED"))
    m = m.reindex(index=classes, columns=classes, fill_value=0)
    m["TOTAL (true)"] = m.sum(axis=1)
    m.loc["TOTAL (predicted)"] = m.sum(axis=0)
    return m.reset_index()


def support_sheet(recs) -> pd.DataFrame:
    keep=["tier","true","pred","severity","n_occurrences","n_events","n_right","n_wrong",
          "wrong_rate_pct","n_streaks","n_dd","final_score","dd1_recovered"]
    d=pd.DataFrame([{k:r[k] for k in keep} for r in _order(recs)])
    d["tier"]=d["tier"].fillna("(class row)")
    return d.rename(columns={c:c.replace("_"," ") for c in d.columns})


def previous_counts(path) -> dict:
    """{(true, pred): count} from the LAST run's workbook, for the Change column.

    read from the file, not from a database, because the workbook IS the record -- it is what
    gets downloaded, annotated and kept. if there is no previous file the column stays blank;
    a first run has nothing to compare against and pretending otherwise would invent a number.
    """
    if not path:
        return {}
    p = pathlib.Path(path)
    if not p.exists():
        print(f"      no previous workbook at {p} -- Change stays blank")
        return {}
    out = {}
    for sheet in (SHEET_BAD, SHEET_OK):
        try:
            d = pd.read_excel(p, sheet_name=sheet)
        except Exception:
            continue
        for _, r in d.iterrows():
            try:
                out[(str(r[COL_TRUE]).strip(), str(r[COL_PRED]).strip())] = int(r[COL_N])
            except Exception:
                pass
    print(f"      Change compared against {p.name}  ({len(out)} pairs)")
    return out


def severity_matrix(classes: list) -> dict:
    """{(true, pred): severity}. from configs/severity_7class.json -- the SAME file the trainer's
    objective and the shap mistake-ranking read, so one number cannot disagree with itself.

    it is read from there, not copied out of a spreadsheet, because a spreadsheet is a snapshot
    and this file is the live setting.
    """
    import json
    if not C.SEVERITY_FILE.exists():
        return {}
    cfg = json.loads(C.SEVERITY_FILE.read_text())
    sev = {k: v for k, v in cfg.get("severity", {}).items() if not k.startswith("_")}
    dflt = float(cfg.get("default", 1))
    return {(a, b): (0.0 if a == b else float(sev.get(f"{a}->{b}", dflt)))
            for a in classes for b in classes}


def weights_used(bundle: dict) -> dict:
    """the CLASS_WEIGHTS this model actually trained under.

    FROM THE BUNDLE, NOT FROM TODAY'S CONFIG. config.CLASS_WEIGHTS is recomputed whenever the
    label shares move -- it was rewritten on 2026-08-08 when L1 went 74.5% -> 86.8% NO_TRADE.
    reading it here would print numbers an older model never saw, and the sheet would quietly
    misattribute every mistake. models trained before this key existed fall back, and say so.
    """
    w = bundle.get("class_weights")
    if w:
        return dict(w)
    fallback = dict(getattr(C, "CLASS_WEIGHTS", {}) or {})
    print(f"      !! this model predates the recorded class_weights. falling back to TODAY's "
          f"config.CLASS_WEIGHTS, which may not be what it trained on: {fallback}")
    return fallback


def cell_table(shap_vals, feats, classes, y_true, y_pred) -> dict:
    """{(true, pred): {"n": rows, feature: mean signed shap toward `pred`}}.

    shap_vals is (rows, features, classes). for each cell we take the column of the PREDICTED
    class and average it over the rows in that cell.
    """
    ci = {c: i for i, c in enumerate(classes)}
    out = {}
    pairs = pd.DataFrame({"t": y_true, "p": y_pred}).groupby(["t", "p"]).groups
    for (t, p), idx in pairs.items():
        if p not in ci:
            continue
        rows = np.asarray(idx)
        v = shap_vals[rows][:, :, ci[p]].mean(axis=0)          # (features,) signed, toward p
        out[(t, p)] = {"n": len(rows), **{f: float(x) for f, x in zip(feats, v)}}
    return out


def sheet_for(pairs: list, cells: dict, feats: list, weights: dict,
              sev: dict, prev: dict) -> pd.DataFrame:
    """one row per (true, predicted) pair -- EVERY pair, whether it happened or not.

    a pair the model never produced still gets a row, with count 0. leaving it out would make
    "never happened" and "not looked at" indistinguishable, and the sheet would change shape from
    run to run.
    """
    rows = []
    for t_, p_ in pairs:
        got = cells.get((t_, p_))
        n_now = got["n"] if got else 0
        # "343-444" -- THIS run's count and the LAST one's, side by side, as written. not the
        # subtraction: seeing both numbers tells you whether a drop is 444->343 or 4->3.
        chg = f"{n_now}-{prev[(t_, p_)]}" if (t_, p_) in prev else ""
        r = {COL_TRUE: t_, COL_PRED: p_,
             COL_SEV: sev.get((t_, p_)),
             COL_N: n_now,
             COL_CHG: chg,
             COL_W: weights.get(t_),
             COL_RAT: ""}                     # the manager fills this in after downloading
        for f in feats:
            r[f] = round(got[f], 6) if got and f in got else (0.0 if got else None)
        rows.append(r)
    return pd.DataFrame(rows)


def paint_diagonal(path: pathlib.Path, classes: list) -> None:
    """green the diagonal of the confusion matrix -- the CORRECT cells.

    done after writing, because pandas has no cell-formatting hook. openpyxl only.
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font
    except ImportError:
        print("      (no openpyxl -- diagonal not coloured)")
        return
    wb = openpyxl.load_workbook(path)
    if SHEET_CM not in wb.sheetnames:
        return
    ws = wb[SHEET_CM]
    green = PatternFill("solid", fgColor="FFA9D08E")
    grey = PatternFill("solid", fgColor="FFF2F2F2")
    for j in range(1, ws.max_column + 1):
        ws.cell(1, j).font = Font(bold=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = 16
    for i, c in enumerate(classes):
        r = 2 + i                       # row 1 is the header
        ws.cell(r, 1).font = Font(bold=True)
        ws.cell(r, 2 + i).fill = green  # column 1 is the TRUE label, so class i sits at 2+i
    last = 2 + len(classes)             # the TOTAL row / column
    for j in range(1, ws.max_column + 1):
        ws.cell(ws.max_row, j).fill = grey
    for i in range(1, ws.max_row + 1):
        ws.cell(i, last).fill = grey
    ws.freeze_panes = "B2"
    wb.save(path)
    print(f"      diagonal coloured green ({len(classes)} correct cells)")


def build(out: pathlib.Path, shap_vals, feats, classes, y_true, y_pred,
          weights: dict, previous=None, times=None, oos=None) -> pathlib.Path:
    """FIVE sheets. every pair, no prose. 7 classes -> 49 rows on each of the first four.

    TWO POPULATIONS, ON PURPOSE:
        the 2 shap sheets + the confusion matrix   -> the TEST slice (shap needs the feature rows)
        drawdowns + losing streaks                 -> the OOS block  (see oos_stream's docstring)
    """
    cells = cell_table(shap_vals, feats, classes, y_true, y_pred)
    sev = severity_matrix(classes)
    prev = previous_counts(previous)
    wrong = [(a, b) for a in classes for b in classes if a != b]
    right = [(a, a) for a in classes]
    print(f"      {len(wrong)} mistake pairs + {len(right)} correct = {len(wrong)+len(right)} rows"
          f"   ({len(cells)} actually occurred)")
    print(f"      {len(feats)} feature columns")

    s_true, s_pred, s_times = oos if oos is not None else (y_true, y_pred, times)
    print(f"      streak sheets walk {len(s_true):,} "
          f"{'OOS' if oos is not None else 'TEST (!! not time-contiguous)'} rows")
    recs = cell_rows(np.asarray(s_true), np.asarray(s_pred), classes, sev, times=s_times)
    cens = sum(1 for r in recs if r["dd1_recovered"] == "no - data ended")
    nodd = sum(1 for r in recs if r["max_dd1"] is None)
    print(f"      {nodd} cells never fell behind at all;  {cens} still in drawdown when the "
          f"data ends (those depths are a FLOOR, not a final figure)")

    sheets = {SHEET_BAD: sheet_for(wrong, cells, feats, weights, sev, prev),
              SHEET_OK:  sheet_for(right, cells, feats, weights, sev, prev),
              SHEET_DD:  dd_sheet(recs),
              SHEET_LS:  ls_sheet(recs),
              SHEET_CM:  cm_sheet(y_true, y_pred, classes)}
    out.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out, engine="openpyxl") as w:   # openpyxl, so we can colour after
        for name, d in sheets.items():
            d.to_excel(w, sheet_name=name[:31], index=False)
    paint_diagonal(out, classes)
    print(f"      wrote {out}   ({len(sheets)} sheets)")
    return out


def from_task(model_task_id: str):
    """fetch the model and its dataset from ClearML, rebuild the SAME test split, and predict.

    it does NOT read a scored_tables artifact. it has the model and the data, so it predicts for
    itself -- which means this step has no sibling to wait for and can be queued alongside shap
    and deepchecks rather than after them.
    """
    from clearml import Dataset, Task
    from trainer.train import load_model_bundle, find_dataset_parquet
    from trainer.objective import three_way_split
    from predict import prepare

    src = Task.get_task(task_id=model_task_id)
    if src is None:
        raise SystemExit(f"no such training task: {model_task_id}")
    if "model" not in src.artifacts:
        raise SystemExit(f"task {model_task_id} has no 'model' artifact -- it failed or never "
                         f"saved one. nothing to analyse.")
    b = load_model_bundle(src.artifacts["model"].get_local_copy())
    feats, classes = list(b["features"]), list(b["label_encoder"].classes_)
    print(f"[1/5] {b.get('model_type','')}   {len(feats)} features, {len(classes)} classes")

    ds = Dataset.get(dataset_id=b["dataset_id"], alias="feature_analysis")
    df = pd.read_parquet(find_dataset_parquet(pathlib.Path(ds.get_local_copy()), b["dataset_id"]))
    ts = pd.to_datetime(df[C.LABEL_TS_COL])

    # THE SPLIT COMES FROM THE BUNDLE, NOT FROM TODAY'S CONFIG -- same reasoning as shap_explain.
    # under bundle_random the test rows are scattered through the whole history, so `test_start`
    # is not a date cut and using it selects nearly everything, including rows the model memorised.
    sp = b.get("split") or {}
    if sp.get("strategy") == "bundle_random":
        _, _, te, _ = three_way_split(
            ts, float(sp.get("val_fraction", C.VAL_FRACTION)),
            float(sp.get("test_fraction", C.TEST_FRACTION)),
            int(sp.get("embargo_sessions", C.EMBARGO_SESSIONS)),
            strategy=sp["strategy"], bundle_minutes=sp.get("bundle_minutes"), seed=sp.get("seed"))
        print(f"[2/5] test split REBUILT from the bundle: {sp['strategy']}, seed {sp.get('seed')}")
    else:
        te = ts > pd.Timestamp(sp["test_start"]) if sp.get("test_start") else \
             pd.Series(True, index=ts.index)
        print(f"[2/5] test split from the bundle's recorded cut {sp.get('test_start')}")
    sub = df[te.to_numpy()]
    print(f"      {len(sub):,} test rows of {len(df):,}")

    y_true = sub[C.LABEL_COL].astype(str).str.strip().to_numpy()
    X = prepare(sub, b)
    y_pred = b["label_encoder"].inverse_transform(b["model"].predict_proba(X).argmax(axis=1))
    # the DRAWDOWN sheets walk the stream IN TIME ORDER, so the timestamps travel too.
    # without them the periods would read 'row 12,345' instead of a date.
    tms = pd.to_datetime(sub[C.LABEL_TS_COL]).to_numpy()
    return b, X, feats, classes, y_true, y_pred, tms


_DONE = {"completed", "failed", "stopped", "closed", "published", "publishing", "unknown"}


def _status_of(row) -> str:
    """ClearML hands back a STATUS ENUM here, not a string.

    str(TaskStatusEnum.completed) is 'TaskStatusEnum.completed', so a plain == "completed" is
    False for a task that finished perfectly -- and a waiter built on that would never stop
    waiting. take .value when it exists and keep only the part after the last dot.
    """
    s = row.get("status") if isinstance(row, dict) else row
    s = getattr(s, "value", s)
    return str(s).rsplit(".", 1)[-1].strip().lower()


def wait_for_siblings(model_task_id: str, project: str, timeout_min: int = 360,
                      poll_sec: int = 60) -> None:
    """block until the other steps of THIS training run have finished.

    WHY WAIT AT ALL. this workbook is the run's summary, so it should be the last thing produced
    -- read it and you know the run is over. it also stops a half-finished picture: someone opens
    the workbook, sees the numbers, and does not realise the backtest is still running.

    MATCHED ON Args/model_task_id, NOT ON THE NAME. two runs of the same version share a name, and
    a name match would silently pair this workbook with a different model's siblings. same rule
    ui/_shared.children_of already follows.

    IT CANNOT DEADLOCK ON ONE AGENT. train.py enqueues this LAST, so a single agent reaches it only
    after the other four are done and the wait returns straight away. with several agents it may
    start early and hold one slot while it waits -- one slot, so the others still progress. the
    timeout is the backstop for the case where a sibling never runs at all (no agent on its queue,
    a base task that was never registered), and it EXPIRES INTO DOING THE WORK rather than into an
    error: a late workbook is worth more than no workbook.
    """
    from clearml import Task
    deadline = time.time() + timeout_min * 60
    time.sleep(15)          # grace: train.py enqueues all five within a second of each other
    seen, waited = 0, False
    while True:
        try:
            rows = Task.query_tasks(
                project_name=project, task_name="^(scored_|shap_|deepchecks_)",
                task_filter={"order_by": ["-last_update"], "page_size": 200, "page": 0},
                additional_return_fields=["id", "name", "status",
                                          "hyperparams.Args.model_task_id"]) or []
        except Exception as exc:
            print(f"      could not list siblings ({exc}) -- not waiting.")
            return
        mine = [r for r in rows
                if str(_val(r.get("hyperparams.Args.model_task_id"))) == model_task_id]
        seen = max(seen, len(mine))
        busy = [r for r in mine if _status_of(r) not in _DONE]
        if not busy:
            if waited:
                print(f"      all {seen} sibling step(s) finished -- starting.")
            elif seen:
                print(f"      {seen} sibling step(s) already finished -- starting.")
            else:
                print(f"      no sibling steps found for {model_task_id} -- starting.")
            return
        if time.time() > deadline:
            print(f"      !! waited {timeout_min} min and {len(busy)} step(s) are still going "
                  f"({', '.join(sorted(r.get('name','?') for r in busy))}). going ahead anyway "
                  f"-- a late workbook beats no workbook.")
            return
        waited = True
        left = int((deadline - time.time()) / 60)
        print(f"      waiting for {len(busy)} of {seen}: "
              f"{', '.join(sorted(str(r.get('name','?')) for r in busy))}   "
              f"({left} min before it gives up)", flush=True)
        time.sleep(poll_sec)


def _val(v):
    """query_tasks returns a hyperparameter as either the bare value or {'value': ...}."""
    return v.get("value") if isinstance(v, dict) else v


def oos_stream(b: dict, dataset_id: str = "", data_path: str = ""):
    """predict on the OOS set and return (y_true, y_pred, times). NO shap.

    THE DRAWDOWN AND LOSING-STREAK SHEETS RUN ON THIS, NOT ON THE TEST SLICE, AND THAT IS A
    CORRECTNESS RULE, NOT A PREFERENCE.
        under bundle_random the test rows are scattered right through 2015-2024. between two
        rows that LOOK consecutive in the test slice sit hundreds of rows the model TRAINED on.
        a run of wrong calls counted across that gap describes an order that never existed, and
        a drawdown walks a timeline with holes in it. both numbers would be fiction.
        the OOS set is one unbroken block (2025-01-01 onward) that starts after every training
        row ends. a streak in it is a real streak, and a period reads as a real date range.

    it costs one predict_proba and nothing else, so it adds a minute -- not an hour like shap.
    """
    from trainer.export_scored_tables import normalise_time_column, preflight
    from predict import prepare

    if data_path:
        src = pathlib.Path(data_path)
        if not src.exists():
            raise SystemExit(f"--oos_data not found: {src}")
        print(f"      OOS from disk: {src}")
    else:
        from clearml import Dataset
        from trainer.train import find_dataset_parquet
        did = dataset_id or C.resolve_dataset_id(C.CLEARML_OOS_DATASET,
                                                 getattr(C, "OOS_DATASET_ID", ""))
        ds = Dataset.get(dataset_id=did, alias="feature_analysis_oos")
        src = find_dataset_parquet(pathlib.Path(ds.get_local_copy()), did)
        print(f"      OOS from ClearML: {C.CLEARML_OOS_DATASET} {did}")

    # the SAME footer-only preflight the scoring step uses. it catches a missing feature and the
    # silent one -- a column that was text in training arriving as numbers -- before any rows load.
    preflight(src, b)
    df = normalise_time_column(pd.read_parquet(src))

    # WHICH truth column. attach_oos_labels writes one per label set (primary_label_L1/_L2/_L3)
    # because truth depends on which set the model trained under. picking the wrong one would look
    # exactly like a bad model, so refuse rather than guess -- same rule as export_scored_tables.
    want = str(b.get("labels_name") or "")
    tagged = sorted(c for c in df.columns if c.startswith(C.LABEL_COL + "_"))
    if want and f"{C.LABEL_COL}_{want}" in df.columns:
        label_col = f"{C.LABEL_COL}_{want}"
    elif C.LABEL_COL in df.columns:
        label_col = C.LABEL_COL
    else:
        raise SystemExit(
            f"the OOS set carries no truth column for this model. it trained on {want or '(none)'}"
            f" and the file has {tagged or 'no label columns at all'}.\n"
            f"  drawdowns and losing streaks need the TRUE label -- they cannot be computed "
            f"without it.\n"
            f"  fix:  final_venv/bin/python scripts/attach_oos_labels.py --sets {want or 'L1'}")
    print(f"      truth column: {label_col}   ({len(df):,} OOS rows)")

    y_true = df[label_col].astype(str).str.strip().to_numpy()
    keep = pd.Series(y_true).isin(list(b["label_encoder"].classes_)).to_numpy()
    if not keep.all():
        print(f"      dropping {int((~keep).sum()):,} row(s) whose label is not one of this "
              f"model's {len(b['label_encoder'].classes_)} classes")
        df, y_true = df[keep], y_true[keep]

    y_pred = b["label_encoder"].inverse_transform(
        b["model"].predict_proba(prepare(df, b)).argmax(axis=1))
    times = pd.to_datetime(df[C.LABEL_TS_COL]).to_numpy()
    # ORDER MATTERS FOR THIS SHEET AND NOTHING ELSE READS IT SORTED, so sort here rather than
    # trusting the file. a drawdown walked out of order is meaningless.
    o = np.argsort(times, kind="stable")
    y_true, y_pred, times = y_true[o], y_pred[o], times[o]
    hit = float((y_true == y_pred).mean()) * 100
    print(f"      scored: {len(y_true):,} rows, {hit:.1f}% correct, "
          f"{pd.Timestamp(times[0]):%d-%m-%Y} to {pd.Timestamp(times[-1]):%d-%m-%Y}")
    return y_true, y_pred, times


def patch_streaks(book: pathlib.Path, b: dict, classes: list, oos) -> pathlib.Path:
    """rewrite ONLY the two streak sheets of a workbook that already exists.

    a workbook built before the oos rule has fiction in those two sheets and hours of correct shap
    in the other three. recomputing everything to fix two sheets would throw away the expensive
    part to redo the cheap one, so this reads the file, replaces exactly those sheets, and puts it
    back with the sheet order and the green diagonal intact.
    """
    if not book.exists():
        raise SystemExit(f"nothing to patch -- no workbook at {book}")
    sheets = {n: pd.read_excel(book, sheet_name=n) for n in pd.ExcelFile(book).sheet_names}
    if SHEET_DD not in sheets or SHEET_LS not in sheets:
        raise SystemExit(f"{book.name} has {list(sheets)} -- no streak sheets to patch.")

    s_true, s_pred, s_times = oos
    recs = cell_rows(np.asarray(s_true), np.asarray(s_pred), classes,
                     severity_matrix(classes), times=s_times)
    old = sheets[SHEET_LS]["Max Losing Streak1"].max()
    sheets[SHEET_DD], sheets[SHEET_LS] = dd_sheet(recs), ls_sheet(recs)
    new = sheets[SHEET_LS]["Max Losing Streak1"].max()
    print(f"      worst losing streak was {old} (test, fiction) -> {new} (oos, real)")

    with pd.ExcelWriter(book, engine="openpyxl") as w:
        for name, d in sheets.items():
            d.to_excel(w, sheet_name=name[:31], index=False)
    paint_diagonal(book, classes)
    print(f"      patched {book}  ({SHEET_DD!r} + {SHEET_LS!r} only, "
          f"the {len(sheets)-2} shap/matrix sheets untouched)")
    return book


def main():
    # CLEARML MUST BE IMPORTED BEFORE parse_args(). it patches argparse at import, and that patch
    # is what feeds a cloned task's Args/* into the parser. importing it afterwards silently loses
    # every override -- the bug that cost the scored-tables step every pipeline run.
    try:
        from clearml import Task           # noqa: F401
    except ImportError:
        Task = None

    ap = argparse.ArgumentParser()
    ap.add_argument("--model_task_id", default="",
                    help="pipeline mode: the training task. fetches the model + dataset itself.")
    ap.add_argument("--bundle", default="", help="local mode: the model joblib")
    ap.add_argument("--scored", default="", help="local mode: a scored_test parquet")
    ap.add_argument("--data", default="", help="local mode: the dataset parquet")
    ap.add_argument("--out", default="feature_analysis_out")
    ap.add_argument("--previous", default="",
                    help="last run's workbook, for the Change column. omit and it takes the "
                         "newest one already in --out.")
    ap.add_argument("--max_rows", type=int,
                    default=int(getattr(C, "FEATURE_ANALYSIS_MAX_ROWS", 0)),
                    help="cap the rows shap runs on. 0 = ALL, which is the setting we run. shap "
                         "is exact per row, so a sample would change only the precision of each "
                         "cell's mean -- but the rare cells are the expensive mistakes and a "
                         "sample leaves them with almost no rows.")
    ap.add_argument("--wait_for_siblings", type=int, default=1,
                    help="1 = run only after this run's other steps finish (the default: this "
                         "workbook is the run's summary). 0 = start immediately.")
    # the drawdown + losing-streak sheets ALWAYS run on OOS. these only say WHERE to find it.
    ap.add_argument("--oos_data", default=str(C.ROOT / "data" / "oos" / "OOS_labelled.parquet"),
                    help="the labelled OOS parquet on disk. pipeline mode ignores this and "
                         "pulls the ClearML dataset instead.")
    ap.add_argument("--oos_dataset_id", default="",
                    help=f"pipeline mode: the OOS ClearML dataset. empty = newest "
                         f"'{C.CLEARML_OOS_DATASET}'.")
    ap.add_argument("--patch_streaks", default="",
                    help="an EXISTING workbook. rewrites only its Drawdowns + Losing Streak "
                         "sheets from OOS and skips shap entirely -- minutes, not hours. for "
                         "workbooks built before the oos rule.")
    ap.add_argument("--allow_test_streaks", action="store_true",
                    help="LAST RESORT. compute the streak sheets on the test slice when no OOS "
                         "set can be reached. under bundle_random those rows are not contiguous "
                         "in time, so the streaks and drawdowns are FICTION. read the numbers as "
                         "broken, not as small.")
    a = ap.parse_args()

    # ---- PIPELINE MODE: become a real ClearML task ---------------------------------------
    # done BEFORE any work so the console log streams live. with no --model_task_id this is the
    # base-task registration run: it exits cleanly, which is exactly what makes it clonable.
    # local mode is --bundle (with --scored --data) or --patch_streaks. anything else is the
    # pipeline, INCLUDING a bare run with no arguments -- that is register_base_trainer creating
    # the template, and it only works if a task actually appears.
    task = None
    if Task is not None and not a.bundle and not a.patch_streaks:
        # openpyxl and shap are DECLARED, not inferred. ClearML records packages by scanning this
        # script's imports, and openpyxl is only ever reached through pandas' excel engine -- so
        # without these lines the agent builds a venv without it and dies on the very last line,
        # after doing all the work. that exact failure already cost us a backtest run.
        Task.add_requirements("openpyxl")
        Task.add_requirements("shap")
        task = Task.init(project_name=C.CLEARML_PROJECT,
                         task_name=getattr(C, "BASE_FEATURE_ANALYSIS_NAME",
                                           "feature_analysis (base)"),
                         task_type=Task.TaskTypes.qc,
                         output_uri=C.feature_analysis_output_uri())
        if not a.model_task_id:
            print("no --model_task_id: base-task registration run. exiting cleanly.")
            task.close()
            return 0
        if a.wait_for_siblings:
            print(f"[0/5] waiting for this run's other steps to finish")
            wait_for_siblings(a.model_task_id, C.CLEARML_PROJECT,
                              timeout_min=int(getattr(C, "FEATURE_ANALYSIS_WAIT_MIN", 360)))

    # ---- PATCH MODE, checked first because it skips everything expensive ----------------
    if a.patch_streaks:
        from trainer.train import load_model_bundle
        if a.bundle:
            b = load_model_bundle(a.bundle)
        elif a.model_task_id:
            from clearml import Task
            src = Task.get_task(task_id=a.model_task_id)
            b = load_model_bundle(src.artifacts["model"].get_local_copy())
        else:
            raise SystemExit("--patch_streaks also needs --bundle (or --model_task_id) -- the "
                             "OOS rows have to be scored by the SAME model that made the book.")
        print(f"[1/2] scoring OOS with {b.get('model_type','')} v{b.get('dataset_version','?')}")
        # patch mode is always run BY HAND on a machine that has the file, so prefer the local
        # copy over a 152 MB download -- even when the model came from a task id.
        _local = pathlib.Path(a.oos_data)
        oos = oos_stream(b, dataset_id=a.oos_dataset_id,
                         data_path=str(_local) if _local.exists() else "")
        print(f"[2/2] patching")
        patch_streaks(pathlib.Path(a.patch_streaks), b, list(b["label_encoder"].classes_), oos)
        return 0

    T0 = time.perf_counter(); _t = {}
    if a.model_task_id:
        b, X, feats, classes, y_true, y_pred, times = from_task(a.model_task_id)
    else:
        if not (a.bundle and a.scored and a.data):
            raise SystemExit("give either --model_task_id, or all three of "
                             "--bundle --scored --data")
        from trainer.train import load_model_bundle
        from predict import prepare
        b = load_model_bundle(a.bundle)
        feats, classes = list(b["features"]), list(b["label_encoder"].classes_)
        print(f"[1/5] {b.get('model_type','')}  {len(feats)} features, {len(classes)} classes")
        sc = pd.read_parquet(a.scored)
        sc = sc[sc["true_label"].notna()]
        if "split" in sc.columns and (sc["split"] == "test").any():
            sc = sc[sc["split"] == "test"]
        print(f"[2/5] {len(sc):,} labelled test rows from {pathlib.Path(a.scored).name}")
        df = pd.read_parquet(a.data)
        if C.LABEL_TS_COL in df.columns:
            df = df.set_index(pd.to_datetime(df[C.LABEL_TS_COL]))
        X = prepare(df.loc[pd.to_datetime(sc[C.LABEL_TS_COL])], b)
        y_true = sc["true_label"].astype(str).str.strip().to_numpy()
        y_pred = sc["predicted_label"].astype(str).str.strip().to_numpy()
        times = pd.to_datetime(sc[C.LABEL_TS_COL]).to_numpy()

    _t["fetch + split + predict"] = time.perf_counter() - T0
    mtype = b.get("model_type", "")

    # ---- OOS, BEFORE SHAP ON PURPOSE ----------------------------------------------------
    # it costs about a minute and shap costs hours. doing it first means a missing OOS set or a
    # wrong label column fails NOW, instead of after a night of shap that then cannot be written.
    print(f"[3/5] OOS for the drawdown + losing-streak sheets")
    _o = time.perf_counter()
    oos = None
    try:
        oos = oos_stream(b, dataset_id=a.oos_dataset_id,
                         data_path="" if a.model_task_id else a.oos_data)
    except SystemExit as exc:
        if not a.allow_test_streaks:
            raise SystemExit(f"{exc}\n\n  the streak sheets need the OOS set and it could not be "
                             f"read. stopping BEFORE shap so nothing is wasted.\n"
                             f"  point --oos_data at the labelled parquet, or pass "
                             f"--allow_test_streaks to accept fiction in those two sheets.")
        print(f"      !! {exc}\n      --allow_test_streaks given: falling back to the TEST slice. "
              f"those streaks are NOT real.")
    _t["oos predict"] = time.perf_counter() - _o
    if a.max_rows and len(X) > a.max_rows:
        # SORTED, so the stream stays in time order -- a drawdown walked out of order
        # is meaningless. sampling thins the stream but never reorders it.
        keep = np.sort(np.random.RandomState(42).choice(len(X), a.max_rows,
                                                       replace=False))
        X, y_true, y_pred, times = X.iloc[keep], y_true[keep], y_pred[keep], times[keep]
        print(f"      sampled {a.max_rows:,} rows for shap (--max_rows 0 for all)")

    print(f"[4/5] shap on {len(X):,} rows x {len(feats)} features x {len(classes)} classes")
    _s = time.perf_counter()
    vals, _ = compute_shap(b["model"], X, mtype, cat_features=b.get("categorical"))
    _t["shap"] = time.perf_counter() - _s

    print(f"[5/5] building the workbook")
    _w = time.perf_counter()
    ver = str(b.get("dataset_version", "") or "").lstrip("v")
    out = pathlib.Path(a.out) / f"feature_analysis_{mtype or 'model'}_v{ver or '?'}.xlsx"
    # DEFAULT TO THE NEWEST FILE ALREADY THERE. picked BEFORE we write this run's file, or it
    # would compare against itself.
    prev = a.previous
    if not prev:
        old = sorted(pathlib.Path(a.out).glob("feature_analysis_*.xlsx"),
                     key=lambda q: q.stat().st_mtime) if pathlib.Path(a.out).exists() else []
        old = [q for q in old if q.resolve() != out.resolve()]
        prev = str(old[-1]) if old else ""
    build(out, np.asarray(vals), feats, classes, y_true, y_pred, weights_used(b),
          previous=prev, times=times, oos=oos)

    _t["workbook"] = time.perf_counter() - _w
    # A MEASURED RATE, PRINTED EVERY RUN. two estimates were wrong today because nothing
    # recorded how long this actually takes. now the next run's cost is arithmetic.
    tot = time.perf_counter() - T0
    print(f"\n      TIMING   total {tot:,.1f}s")
    for k, v in _t.items():
        print(f"         {k:<26}{v:>9,.1f}s{v/tot*100:>7.0f}%")
    print(f"         {'':<26}{'':>9} {len(X):,} rows -> {_t['shap']/len(X)*1000:.2f} ms/row (shap)")
    full = 274605
    print(f"         all {full:,} test rows would be about "
          f"{_t['shap']/len(X)*full/60:,.0f} min of shap")

    # UPLOAD ONLY IN PIPELINE MODE. a local run leaves the file on disk; a queued run must put it
    # somewhere that survives the agent, or the whole point of the step is lost.
    if task is not None:
        name = getattr(C, "FEATURE_ANALYSIS_ARTIFACT", "summary_file")
        try:
            # wait_on_upload, so the task cannot close while the file is still going up. an
            # artifact that is still uploading when the process exits is silently lost.
            task.upload_artifact(name, str(out), wait_on_upload=True)
            _lbl = b.get("labels_name")
            task.add_tags([mtype or "?", "feature_analysis", f"v{ver or '?'}"]
                          + ([str(_lbl)] if _lbl else []))
            print(f"      uploaded as artifact '{name}'  ->  {C.feature_analysis_output_uri()}")
            print(f"      download it from the ARTIFACTS tab of this task in the ClearML UI.")
        except Exception as exc:
            # REPORT ONLY. the workbook is on the agent's disk either way, and a failed upload
            # must not mark the whole analysis as failed.
            print(f"      !! could not upload: {exc}")
        task.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
